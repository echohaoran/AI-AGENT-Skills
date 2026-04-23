"""
Excel file handler with cell-level text extraction.
"""

from typing import Optional

from scripts.handlers.base import BaseHandler


class ExcelHandler(BaseHandler):
    """Handles Excel file reading via openpyxl."""

    def read(self, file_path: str) -> Optional[str]:
        """Extract all cell values from Excel."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                parts.append(f"[Sheet: {sheet_name}]")
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(c) if c is not None else "" for c in row]
                    parts.append("\t".join(row_values))
                parts.append("")
            return "\n".join(parts)
        except Exception:
            return None
